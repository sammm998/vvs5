"""Exports: Excel, CSV, JSON, analysis report and marked PDF are produced from the frozen artifacts."""
from __future__ import annotations

import csv
import io
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HEADERS = ["Beteckning", "DN", "Antal fysiska rör", "Horisontellt m", "Vertikalt m", "Totalt m", "Tvetydigt m", "Varav i skrafferat område m", "Stigare (antal)", "Status"]


def _rows(result_dir: str, floor_height: float | None = None, include_hatched: bool = False) -> list[dict]:
    with open(os.path.join(result_dir, "quantities.json"), "r", encoding="utf-8") as fh:
        rows = json.load(fh)["rows"]
    for r in rows:
        # pipe drawn inside hatched areas is measured but stays out of the total unless the takeoff includes it
        if include_hatched:
            hatched = float(r.get("in_hatched_area_m", 0.0) or 0.0)
            r["confirmed_horizontal_m"] = round(r["confirmed_horizontal_m"] + hatched, 3)
            if r["vertical_m"] != "UNKNOWN":
                r["confirmed_total_m"] = round(r["confirmed_horizontal_m"] + float(r["vertical_m"]), 3)
            else:
                r["confirmed_total_m"] = r["confirmed_horizontal_m"]
        if floor_height:
            # vertical metres from the user's floor height: every riser counts one floor height
            risers = int(r.get("riser_count", 0) or 0)
            if risers > 0:
                known = 0.0 if r["vertical_m"] == "UNKNOWN" else float(r["vertical_m"])
                r["vertical_m"] = round(known + risers * floor_height, 3)
                r["confirmed_total_m"] = round(r["confirmed_horizontal_m"] + r["vertical_m"], 3)
    return rows


def _fmt(v):
    return v if v != "UNKNOWN" else "OKÄNT"


def to_xlsx(result_dir: str, floor_height: float | None = None, include_hatched: bool = False) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mängder"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in _rows(result_dir, floor_height, include_hatched):
        ws.append([r["designation"], r["dn"] if r["dn"] is not None else "?", r["physical_pipe_count"], round(r["confirmed_horizontal_m"], 2),
                   _fmt(r["vertical_m"]) if r["vertical_m"] == "UNKNOWN" else round(r["vertical_m"], 2), round(r["confirmed_total_m"], 2),
                   round(r["ambiguous_m"], 2), round(r.get("in_hatched_area_m", 0.0), 2), r.get("riser_count", 0), r["state"]])
    for i, _ in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(result_dir: str, floor_height: float | None = None, include_hatched: bool = False) -> str:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(HEADERS)
    for r in _rows(result_dir, floor_height, include_hatched):
        w.writerow([r["designation"], r["dn"] if r["dn"] is not None else "?", r["physical_pipe_count"], f"{r['confirmed_horizontal_m']:.2f}",
                    _fmt(r["vertical_m"]) if r["vertical_m"] == "UNKNOWN" else f"{r['vertical_m']:.2f}", f"{r['confirmed_total_m']:.2f}",
                    f"{r['ambiguous_m']:.2f}", f"{r.get('in_hatched_area_m', 0.0):.2f}", r.get("riser_count", 0), r["state"]])
    return buf.getvalue()
